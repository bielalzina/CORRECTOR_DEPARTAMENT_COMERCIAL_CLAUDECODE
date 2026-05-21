"""Gestió de les dades de referència (COMPRES i VENDES reals) per tasca i grup."""

import io
import json
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill  # Alignment eliminat (ja no s'usa)
from openpyxl.utils import get_column_letter

TASQUES_DIR = Path(__file__).parent.parent / "data" / "tasques"

COLS_COMPRES = [
    "Expedient", "R_EMPRESA_C", "R_PROVEEDOR_C",
    "R_FECHA_EMISION_C", "R_NUMERO_CP", "R_NUMERO_CA", "R_NUMERO_CF", "R_IMPORTE_C",
]
COLS_VENDES = [
    "Expedient", "R_EMPRESA_V",
    "R_FECHA_EMISION_VP", "R_NUMERO_VP", "R_CLIENTE_V", "R_IMPORTE_V",
    "R_FECHA_MAX_FACTURACION_V",
]

PROVEÏDORS = ["ROCALLA SA", "ALUBIX SL"]
CLIENTS = ["BIGCORP SL", "COMERCIAL CALCO SA"]


def _referencia_path(num_tasca: str, grup: str) -> Path:
    return TASQUES_DIR / num_tasca / grup / "referencia.json"


def load_referencia(num_tasca: str, grup: str) -> dict:
    """Retorna el dict {expedient: {compres: [...], vendes: [...]}} o {} si no existeix."""
    path = _referencia_path(num_tasca, grup)
    if not path.exists():
        return {}
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def save_referencia(num_tasca: str, grup: str, data: dict) -> None:
    path = _referencia_path(num_tasca, grup)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def get_referencia_alumne(num_tasca: str, grup: str, expedient: str) -> Optional[dict]:
    return load_referencia(num_tasca, grup).get(str(expedient))


def generar_plantilla(num_tasca: str, grup: str, alumnes: list[dict]) -> bytes:
    """Genera el fitxer xlsx plantilla buit (capçaleres a la fila 1, sense dades pre-emplenades)."""
    wb = Workbook()

    for nom_full, cols in [("COMPRES", COLS_COMPRES), ("VENDES", COLS_VENDES)]:
        ws = wb.create_sheet(nom_full)
        fill_cap = PatternFill("solid", fgColor="1F4E79")
        font_cap = Font(bold=True, color="FFFFFF")
        for j, col in enumerate(cols, 1):
            c = ws.cell(1, j, col)
            c.fill = fill_cap
            c.font = font_cap
        for j, amplada in enumerate(
            [10, 30, 16, 18, 18, 18, 18, 12] if nom_full == "COMPRES"
            else [10, 30, 18, 18, 20, 12, 22],
            1,
        ):
            ws.column_dimensions[get_column_letter(j)].width = amplada

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def processar_plantilla(contingut: bytes, num_tasca: str, grup: str) -> tuple[dict, list[str]]:
    """
    Llegeix el xlsx emplenat pel professor i retorna (dades, errors).
    dades = {expedient: {compres: [...], vendes: [...]}}
    """
    errors: list[str] = []

    try:
        xls = pd.ExcelFile(io.BytesIO(contingut))
    except Exception as e:
        return {}, [f"No s'ha pogut llegir el fitxer: {e}"]

    for full in ["COMPRES", "VENDES"]:
        if full not in xls.sheet_names:
            errors.append(f"Falta el full '{full}' al fitxer.")
    if errors:
        return {}, errors

    df_c = pd.read_excel(io.BytesIO(contingut), sheet_name="COMPRES", header=0, dtype=str)
    df_v = pd.read_excel(io.BytesIO(contingut), sheet_name="VENDES",  header=0, dtype=str)

    df_c.columns = [str(c).strip() for c in df_c.columns]
    df_v.columns = [str(c).strip() for c in df_v.columns]

    errors += _validar_full(df_c, COLS_COMPRES, "COMPRES")
    errors += _validar_full(df_v, COLS_VENDES,  "VENDES")
    if errors:
        return {}, errors

    dades: dict = {}

    # Processar COMPRES
    for _, row in df_c.iterrows():
        exp = str(row["Expedient"]).strip().split(".")[0]  # eliminar decimals si és número
        if not exp:
            continue
        if row.get("R_PROVEEDOR_C", "").strip() not in PROVEÏDORS:
            errors.append(
                f"COMPRES fila {_ + 2}: proveïdor '{row.get('R_PROVEEDOR_C')}' no vàlid. "
                f"Ha de ser: {', '.join(PROVEÏDORS)}"
            )
            continue
        registre = {
            "R_EMPRESA_C":       str(row.get("R_EMPRESA_C", "")).strip(),
            "R_PROVEEDOR_C":     str(row.get("R_PROVEEDOR_C", "")).strip(),
            "R_FECHA_EMISION_C": _normalitzar_data(row.get("R_FECHA_EMISION_C", "")),
            "R_NUMERO_CP":       str(row.get("R_NUMERO_CP", "")).strip(),
            "R_NUMERO_CA":       str(row.get("R_NUMERO_CA", "")).strip(),
            "R_NUMERO_CF":       str(row.get("R_NUMERO_CF", "")).strip(),
            "R_IMPORTE_C":       _normalitzar_import(row.get("R_IMPORTE_C", "")),
        }
        err = _validar_registre_compra(registre, exp)
        if err:
            errors.extend(err)
        else:
            dades.setdefault(exp, {"compres": [], "vendes": []})["compres"].append(registre)

    # Processar VENDES
    for _, row in df_v.iterrows():
        exp = str(row["Expedient"]).strip().split(".")[0]
        if not exp:
            continue
        if row.get("R_CLIENTE_V", "").strip() not in CLIENTS:
            errors.append(
                f"VENDES fila {_ + 2}: client '{row.get('R_CLIENTE_V')}' no vàlid. "
                f"Ha de ser: {', '.join(CLIENTS)}"
            )
            continue
        registre = {
            "R_EMPRESA_V":              str(row.get("R_EMPRESA_V", "")).strip(),
            "R_FECHA_EMISION_VP":       _normalitzar_data(row.get("R_FECHA_EMISION_VP", "")),
            "R_NUMERO_VP":              str(row.get("R_NUMERO_VP", "")).strip(),
            "R_CLIENTE_V":              str(row.get("R_CLIENTE_V", "")).strip(),
            "R_IMPORTE_V":              _normalitzar_import(row.get("R_IMPORTE_V", "")),
            "R_FECHA_MAX_FACTURACION_V": _normalitzar_data(row.get("R_FECHA_MAX_FACTURACION_V", "")),
        }
        err = _validar_registre_venda(registre, exp)
        if err:
            errors.extend(err)
        else:
            dades.setdefault(exp, {"compres": [], "vendes": []})["vendes"].append(registre)

    return dades, errors


def _validar_full(df: pd.DataFrame, cols: list[str], nom: str) -> list[str]:
    errors = []
    for col in cols:
        if col not in df.columns:
            errors.append(f"Full '{nom}': falta la columna '{col}'.")
    return errors


def _validar_registre_compra(r: dict, exp: str) -> list[str]:
    errors = []
    camps_obligatoris = ["R_FECHA_EMISION_C", "R_NUMERO_CP", "R_NUMERO_CA", "R_NUMERO_CF", "R_IMPORTE_C"]
    for camp in camps_obligatoris:
        val = r.get(camp)
        if not val or val in ("nan", "None", ""):
            errors.append(f"COMPRES exp.{exp}: falta '{camp}'.")
    if r.get("R_IMPORTE_C") is not None and not isinstance(r["R_IMPORTE_C"], (int, float)):
        errors.append(f"COMPRES exp.{exp}: 'R_IMPORTE_C' ha de ser un número.")
    return errors


def _validar_registre_venda(r: dict, exp: str) -> list[str]:
    errors = []
    camps_obligatoris = ["R_FECHA_EMISION_VP", "R_NUMERO_VP", "R_IMPORTE_V", "R_FECHA_MAX_FACTURACION_V"]
    for camp in camps_obligatoris:
        val = r.get(camp)
        if not val or val in ("nan", "None", ""):
            errors.append(f"VENDES exp.{exp}: falta '{camp}'.")
    return errors


def _normalitzar_data(valor) -> str:
    """Converteix diversos formats de data a ISO yyyy-mm-dd."""
    if not valor or str(valor).strip() in ("nan", "None", ""):
        return ""
    s = str(valor).strip()
    # Pandas pot llegir-ho com a datetime
    try:
        import pandas as pd
        ts = pd.to_datetime(s, dayfirst=True)
        return ts.strftime("%Y-%m-%d")
    except Exception:
        return s


def _normalitzar_import(valor) -> Optional[float]:
    if not valor or str(valor).strip() in ("nan", "None", ""):
        return None
    try:
        return float(str(valor).replace(",", "."))
    except ValueError:
        return None


def referencia_completa(num_tasca: str, grup: str, alumnes: list[dict]) -> bool:
    """Retorna True si hi ha dades de referència per a tots els alumnes actius."""
    ref = load_referencia(num_tasca, grup)
    return all(str(a["expedient"]) in ref for a in alumnes)
